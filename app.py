import streamlit as st
import pandas as pd
import io
from database import DialogDatabase
from analyzer import DialogAnalyzer
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

st.set_page_config(page_title="Анализ диалогов", layout="wide")

st.markdown("""
<style>
    .main-title { font-size: 2rem; text-align: center; margin-bottom: 1.5rem; }
    .problem-card { background: #ffebee; padding: 1rem; border-radius: 10px; margin: 0.5rem 0; }
    .similar-card { background: #f5f5f5; padding: 1rem; border-radius: 10px; margin: 0.5rem 0; }
    div[data-testid="stMetricValue"] { color: #000000 !important; }

    /* Меняем текст кнопки Upload */
    div[data-testid="stFileUploader"] button div p {
        font-size: 0;
    }

    div[data-testid="stFileUploader"] button div p::after {
        content: "Загрузить";
        font-size: 16px;
    }

    .stats-box {
        background: #f8f9fa;
        border: 1px solid #e6e6e6;
        border-radius: 12px;
        padding: 12px 16px;
        margin-bottom: 10px;
    }

    .stats-row {
        display: flex;
        justify-content: space-between;
        align-items: center;
        gap: 10px;
    }

    .stats-name {
        font-size: 16px;
        font-weight: 500;
        color: #1f2937;
    }

    .stats-right {
        display: flex;
        align-items: center;
        gap: 10px;
    }

    .stats-count {
        background: #2563eb;
        color: white;
        padding: 4px 10px;
        border-radius: 999px;
        font-size: 14px;
        font-weight: 600;
    }

    .stats-percent {
        color: #6b7280;
        font-size: 14px;
        min-width: 55px;
        text-align: right;
    }

    .similar-card {
        background: #f5f7fa;
        border: 1px solid #e0e0e0;
        border-radius: 12px;
        padding: 14px 18px;
        margin-bottom: 12px;
        font-size: 16px;
    }
    
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">Анализ диалогов</div>', unsafe_allow_html=True)

@st.cache_resource
def init():
    return DialogDatabase(), DialogAnalyzer()

db, analyzer = init()

import io
import pandas as pd
import streamlit as st


def validate_csv_file(file) -> tuple[bool, str, pd.DataFrame | None]:
    #проверка размера файла
    if file.size == 0:
        return False, "Файл пустой", None

    if file.size > 50 * 1024 * 1024:  # 50 MB лимит
        return False, "Файл слишком большой (>50 MB)", None

    #попытка чтения с разными кодировками
    df = None
    encodings = ['utf-8', 'cp1251', 'utf-8-sig', 'latin-1']

    for enc in encodings:
        try:
            content = file.getvalue().decode(enc)
            df = pd.read_csv(io.StringIO(content))
            break
        except UnicodeDecodeError:
            continue
        except pd.errors.EmptyDataError:
            return False, "Файл не содержит данных", None
        except pd.errors.ParserError as e:
            return False, f"Ошибка парсинга CSV: {str(e)}", None
    else:
        return False, "Не удалось определить кодировку файла", None

    #проверка обязательной колонки
    if 'dialog' not in df.columns:
        available = ', '.join(df.columns.tolist())
        return False, f"Отсутствует колонка 'dialog'. Найдено: [{available}]", None

    #проверка на пустые диалоги
    if df['dialog'].dropna().empty:
        return False, "Колонка 'dialog' не содержит данных", None

    #полностью убираем пустые строки
    df = df.dropna(subset=['dialog']).reset_index(drop=True)

    if df.empty:
        return False, "После очистки файл не содержит валидных диалогов", None

    return True, f"OK: {len(df)} диалогов", df


with st.sidebar:
    st.header("Загрузка файлов")

    uploaded_files = st.file_uploader(
        "CSV с колонкой 'dialog'",
        type=['csv'],
        accept_multiple_files=True,
        help="Можно выбрать несколько файлов. Каждый файл должен содержать колонку 'dialog'."
    )

    if uploaded_files:
        all_dfs = []
        errors = []

        progress_bar = st.progress(0)
        status_text = st.empty()

        for idx, file in enumerate(uploaded_files):
            status_text.text(f"Обработка: {file.name}...")

            is_valid, message, df = validate_csv_file(file)

            if is_valid:
                all_dfs.append(df)
                st.success(f"{file.name}: {message}")
            else:
                errors.append(f"{file.name}: {message}")
                st.error(f"{file.name}: {message}")

            # Обновляем прогресс
            progress_bar.progress((idx + 1) / len(uploaded_files))

        progress_bar.empty()
        status_text.empty()

        #вывод сводки по ошибкам
        if errors:
            with st.expander(f"Ошибки в {len(errors)} файлах", expanded=False):
                for err in errors:
                    st.caption(err)

        #продолжаем, если есть 1 или более корректных файлов
        if all_dfs:
            # Объединяем все валидные DataFrame
            df_combined = pd.concat(all_dfs, ignore_index=True)

            st.divider()
            st.info(f"Загружено: **{len(df_combined)} диалогов** из {len(all_dfs)} файлов")

            #показ данных
            with st.expander("Просмотр данных", expanded=False):
                st.dataframe(df_combined.head(100), use_container_width=True)

            #кнопка анализа
            if st.button("Анализировать", type="primary", use_container_width=True):
                with st.spinner("Идет анализ диалогов..."):
                    try:
                        #конвертируем в CSV для load_dialogs
                        csv_buffer = io.StringIO(df_combined.to_csv(index=False))
                        loaded_count = db.load_dialogs(csv_buffer)

                        #анализ через analyzer
                        results = [analyzer.analyze_dialog(d) for d in df_combined['dialog']]

                        #добавляем результаты в DataFrame
                        df_combined['тема'] = [r['topic'] for r in results]
                        df_combined['эмоция'] = [r['emotion'] for r in results]
                        df_combined['проблемный'] = [r['is_problem'] for r in results]

                        #сохраняем в session_state
                        st.session_state.df = df_combined
                        st.session_state.loaded_count = loaded_count

                        st.success(f"Анализ завершён! Обработано: {len(df_combined)} диалогов")
                        st.rerun()

                    except Exception as e:
                        st.error(f"Критическая ошибка при анализе: {str(e)}")
                        st.exception(e)  # Покажет стектрейс в режиме разработки
        else:
            st.warning("Нет валидных файлов для анализа. Исправьте ошибки выше и попробуйте снова.")

if 'df' in st.session_state:
    df = st.session_state.df
    
    c1, c2, c3, c4 = st.columns(4)
    with c1: st.metric("Всего", len(df))
    with c2: st.metric("Проблемных", df['проблемный'].sum())
    with c3: st.metric("Негативных", (df['эмоция'] == 'негативный').sum())
    with c4: st.metric("Позитивных", (df['эмоция'] == 'позитивный').sum())

    #экспорт результатов

    export_df = df.copy()

    export_df['проблемный'] = export_df['проблемный'].map({
        True: 'Да',
        False: 'Нет'
    })

    export_df = export_df.rename(columns={
        "dialog": "Диалог",
        "тема": "Тема",
        "эмоция": "Эмоция",
        "проблемный": "Проблемный"
    })

    #CSV
    csv_result = export_df.to_csv(
        index=False,
        sep=";",
        encoding="utf-8-sig"
    )

    #Excel
    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        export_df.to_excel(
            writer,
            index=False,
            sheet_name="Результаты анализа"
        )

        worksheet = writer.sheets["Результаты анализа"]

        worksheet.column_dimensions["A"].width = 80
        worksheet.column_dimensions["B"].width = 22
        worksheet.column_dimensions["C"].width = 18
        worksheet.column_dimensions["D"].width = 15

        header_fill = PatternFill(
            start_color="D9EAF7",
            end_color="D9EAF7",
            fill_type="solid"
        )

        thin_border = Border(
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF")
        )

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )

        worksheet.row_dimensions[1].height = 22

        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 28

    #две кнопки рядом
    download_col1, download_col2 = st.columns(2)

    with download_col1:
        st.download_button(
            label="Скачать результаты CSV",
            data=csv_result.encode("utf-8-sig"),
            file_name="analysis_results.csv",
            mime="text/csv",
            use_container_width=True
        )

    with download_col2:
        st.download_button(
            label="Скачать результаты Excel",
            data=excel_buffer.getvalue(),
            file_name="analysis_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    
    tab1, tab2, tab3, tab4 = st.tabs(["Диалоги", "Проблемные", "Статистика", "Поиск"])
    
    with tab1:
        display = df[['dialog', 'тема', 'эмоция']].copy()
        display.index = range(1, len(display) + 1)
        st.dataframe(display, use_container_width=True, height=500)

    with tab2:
        problems = df[df['проблемный'] == True]

        if len(problems) > 0:
            for _, row in problems.iterrows():
                if row['эмоция'] == 'нейтральный':
                    bg_color = '#fff9c4'
                    border_color = '#fbc02d'
                elif row['эмоция'] == 'негативный':
                    bg_color = '#ffebee'
                    border_color = '#e57373'

                st.markdown(f"""
                <div class="problem-card" style="
                    background: {bg_color};
                    border-left: 4px solid {border_color};
                    padding: 1rem;
                    border-radius: 8px;
                    margin: 0.5rem 0;
                ">
                    <strong>{row['тема']}</strong> | 
                    <span style="color: {border_color}; font-weight: 600;">
                        {row['эмоция']}
                    </span><br>
                    {row['dialog'][:200]}...
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("Нет проблемных обращений")

    with tab3:
        topic_counts = df['тема'].value_counts()
        emotion_counts = df['эмоция'].value_counts()

        #выбор типа визуализации
        chart_type = st.selectbox(
            "Тип диаграммы",
            ["Гистограмма", "Круговая", "Лепестковая"],
            key="chart_type_selector"
        )

        col1, col2 = st.columns(2)
        with col1:
            st.markdown("### Распределение по темам")

            if chart_type == "Гистограмма":
                st.bar_chart(topic_counts)
            elif chart_type == "Круговая":
                import plotly.express as px

                fig = px.pie(
                    values=topic_counts.values,
                    names=topic_counts.index,
                    hole=0.4
                )
                fig.update_layout(margin=dict(t=0, b=0, l=0, r=0))
                st.plotly_chart(fig, use_container_width=True)
            elif chart_type == "Лепестковая":
                import plotly.graph_objects as go

                fig = go.Figure()
                values = topic_counts.values.tolist() + [topic_counts.values[0]]
                labels = topic_counts.index.tolist() + [topic_counts.index[0]]
                fig.add_trace(go.Scatterpolar(
                    r=values,
                    theta=labels,
                    fill='toself',
                    name='Количество',
                    line_color='#1f77b4',
                    fillcolor='rgba(31, 119, 180, 0.3)'
                ))
                fig.update_layout(
                    polar=dict(
                        radialaxis=dict(visible=True, range=[0, max(topic_counts.values) * 1.1])
                    ),
                    showlegend=False,
                    margin=dict(t=20, b=20, l=20, r=20),
                    height=400
                )
                st.plotly_chart(fig, use_container_width=True)

        with col2:
            st.markdown("### Распределение по эмоциям")

            if chart_type == "Гистограмма":
                st.bar_chart(emotion_counts)
            elif chart_type == "Круговая":
                import plotly.express as px

                fig = px.pie(
                    values=emotion_counts.values,
                    names=emotion_counts.index,
                    hole=0.4
                )
                fig.update_layout(margin=dict(t=0, b=0, l=0, r=0))
                st.plotly_chart(fig, use_container_width=True)
            elif chart_type == "Лепестковая":
                import plotly.graph_objects as go

                fig = go.Figure()
                values = emotion_counts.values.tolist() + [emotion_counts.values[0]]
                labels = emotion_counts.index.tolist() + [emotion_counts.index[0]]
                fig.add_trace(go.Scatterpolar(
                    r=values,
                    theta=labels,
                    fill='toself',
                    name='Количество',
                    line_color='#2ca02c',
                    fillcolor='rgba(44, 160, 44, 0.3)'
                ))
                fig.update_layout(
                    polar=dict(
                        radialaxis=dict(visible=True, range=[0, max(emotion_counts.values) * 1.1])
                    ),
                    showlegend=False,
                    margin=dict(t=20, b=20, l=20, r=20),
                    height=400
                )
                st.plotly_chart(fig, use_container_width=True)
    with tab4:
        st.subheader("Поиск похожих обращений")

        query = st.text_input(
            "Введите текст обращения",
            placeholder="Например: заказ не пришёл, товар сломался, деньги списали..."
        )

        if query:
            with st.spinner("Ищем похожие обращения..."):

                similar = db.find_similar(
                    query,
                    top_k=5
                )

                if similar:
                    for i, (text, score) in enumerate(similar, 1):
                        st.markdown(f"""
                        <div class="similar-card">
                            <strong>{i} — совпадение {score}%</strong><br>
                            {text}
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.info("Похожие обращения не найдены")

else:
    st.info("Загрузите CSV файл")
